
from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'Salary Register'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    ws.append(["                                                                                                                                                              Qatar Polymer Industrial Company","","","","","","","","","","","","","","","","","","","","","",""])
    ws.append(["                                                                                                                                                                                  Salary Register","","","","","","","","","","","","","","","","","","","","","",""])
    ws.append(["                                                                                                                                                                       "+ args.from_date + "" + "  to  " + "" + args.to_date,"","","","","","","","","","","","","","","","","","","","",""])
    ws.append(["S.No","Emp.No","Name","Zone","Nationality","Grade","Basic ","Days","Earned Basic","Overtime","WOT","HOT","TOT",
    "HRA","CA","Mobile Allowance","Monthly Allowance","Medical Allowance","Food Allowance","Production Incentive","Arrear","Vacation","Air Ticket Allowance","Other Allowance","Gross Pay","Loan","Advance","Other Deduction","Total Deduction","Net Pay"])
    if args.department:
        salary_slips = frappe.get_all("Salary Slip",{'start_date':args.from_date,'end_date':args.to_date,"department":args.department,"docstatus":0},['*']) 
    elif args.grade:
        salary_slips = frappe.get_all("Salary Slip",{'start_date':args.from_date,'end_date':args.to_date,"grade":args.grade,"docstatus":0},['*']) 
    else:
        salary_slips = frappe.get_all("Salary Slip",{'start_date':args.from_date,'end_date':args.to_date,"docstatus":0},['*'])  

    i=1
    basicss = 0
    paydays = 0
    earneddays = 0
    over_t = 0
    w_ot = 0
    h_ot = 0
    tot = 0
    h_r_a = 0
    c_a = 0
    m_a = 0
    m_a_a = 0
    f_a = 0
    p_i = 0
    a_r_r = 0
    o_a = 0
    grs_pay = 0
    l_o = 0
    a_d = 0
    o_t_h = 0
    tot_ded = 0
    n_pay = 0
    vac = 0
    ma_a_1 = 0
    a_t_a = 0

    for ss in salary_slips:
        basic = frappe.get_value('Salary Detail',{'salary_component':'Basic','parent':ss.name },['amount']) or 0
        emp = frappe.get_doc('Employee',{"employee":ss.employee},['*'])
        oa = frappe.get_value('Salary Detail',{'salary_component':'Other Allowance','parent':ss.name},['amount']) or 0
        ata = frappe.get_value('Salary Detail',{'salary_component':'Air Ticket','parent':ss.name},['amount']) or 0
        va = frappe.get_value('Salary Detail',{'salary_component':'Vacation','parent':ss.name},['amount']) or 0
        o = frappe.get_value('Salary Detail',{'salary_component':'Others','parent':ss.name},['amount']) or 0
        ma = frappe.get_value('Salary Detail',{'salary_component':'Mobile Allowance','parent':ss.name},['amount']) or 0
        arr = frappe.get_value('Salary Detail',{'salary_component':'Arrear','parent':ss.name},['amount']) or 0
        fa = frappe.get_value('Salary Detail',{'salary_component':'Food Allowance','parent':ss.name},['amount']) or 0
        ot = frappe.get_value('Salary Detail',{'salary_component':'Overtime','parent':ss.name },['amount']) or 0
        maa = frappe.get_value('Salary Detail',{'salary_component':'Monthly Allowance','parent':ss.name},['amount']) or 0
        lo = frappe.get_value('Salary Detail',{'salary_component':'Loan','parent':ss.name},['amount']) or 0
        pi = frappe.get_value('Salary Detail',{'salary_component':'Production Incentive','parent':ss.name},['amount']) or 0
        ad = frappe.get_value('Salary Detail',{'salary_component':'Advance','parent':ss.name},['amount']) or 0
        oth = frappe.get_value('Salary Detail',{'salary_component':'Others','parent':ss.name},['amount']) or 0
        hra = frappe.get_value('Salary Detail',{'salary_component':"House Rent Allowance",'parent':ss.name},["amount"]) or 0
        ca = frappe.get_value('Salary Detail',{'salary_component':"Conveyance Allowance",'parent':ss.name},["amount"]) or 0
        wot = frappe.get_value('Salary Detail',{'salary_component':"Weekend OT",'parent':ss.name},["amount"]) or 0
        hot = frappe.get_value('Salary Detail',{'salary_component':"Holiday OT",'parent':ss.name},["amount"]) or 0
        maa_1 = frappe.get_value('Salary Detail',{'salary_component':"Medical Allowance",'parent':ss.name},["amount"]) or 0


        total = round(ot+wot+hot)
        basicss += round(emp.basic)
        paydays += round(ss.payment_days)
        earneddays += round(basic)
        over_t += round(ss.overtime)
        w_ot += round(ss.weekend_ot)
        h_ot += round(ss.holiday_ot)
        tot += round(total)
        h_r_a +=round(hra)
        c_a +=round(ca)
        m_a +=round(ma)
        m_a_a +=round(maa)
        f_a += round(fa)
        p_i += round(pi)
        a_r_r += round(arr)
        a_t_a += round(ata)
        o_a += round(oa)
        grs_pay +=round(ss.gross_pay)
        l_o += round(lo)
        a_d += round(ad)
        o_t_h += round(oth)
        tot_ded += round(ss.total_deduction)
        n_pay += round(ss.net_pay)
        vac += round(va)
        ma_a_1 += round(maa_1)
        
        # ws.append([i,emp.employee,emp.employee_name,emp.work_location,emp.nationality,emp.grade,emp.basic,hra,ca,ss.payment_days,basic,ss.overtime,ss.weekend_ot,ss.holiday_ot,o,pi,maa,arr,ma,fa,ss.gross_pay,lo,ad,oth,ss.total_deduction,ss.net_pay,""])
        ws.append([i,emp.employee,emp.employee_name,emp.work_location,emp.nationality,emp.grade,emp.basic,ss.payment_days,round(basic,0),ss.overtime,ss.weekend_ot,ss.holiday_ot,total,hra,ca,ma,maa,maa_1,round(fa,0),pi,arr,round(va),round(ata,0),oa,ss.gross_pay,lo,ad,oth,round(ss.total_deduction,0),round(ss.net_pay,0),""])

        i=1+i
    ws.append(["Total","","","","","", "{:,}".format(basicss), "{:,}".format(paydays), "{:,}".format(earneddays), "{:,}".format(over_t), "{:,}".format(w_ot), "{:,}".format(h_ot), "{:,}".format(tot), "{:,}".format(h_r_a), "{:,}".format(c_a), "{:,}".format(m_a), "{:,}".format(m_a_a), "{:,}".format(ma_a_1), "{:,}".format(f_a), "{:,}".format(p_i), "{:,}".format(a_r_r), "{:,}".format(vac), "{:,}".format(a_t_a), "{:,}".format(o_a), "{:,}".format(grs_pay), "{:,}".format(l_o), "{:,}".format(a_d), "{:,}".format(o_t_h), "{:,}".format(tot_ded), "{:,}".format(n_pay)])
    # numbers = "{:,}".format(5000000)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=30)
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=30)
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=30)
    
    
    
    align_center = Alignment(horizontal='right')
    # vertical = Alignment(textRotation=180,vertical='center',horizontal='center')
        
    if ss.grade == "Factory Staff":
        for cell in ws["15:15"]:
            cell.alignment = align_center
    elif ss.grade == "Office Staff":
        for cell in ws["21:21"]:
            cell.alignment = align_center
    else:
        for cell in ws["290:290"]:
            cell.alignment = align_center
        
    # for cell in ws["2:2"]:
    #     cell.alignment = vertical

    bold_font = Font(bold=True)
    for cell in ws["4:4"]:
        cell.font = bold_font
        
    
    if ss.grade == "Factory Staff":
        for cell in ws["15:15"]:
            cell.font = bold_font
    elif ss.grade == "Office Staff":
        for cell in ws["21:21"]:
            cell.font = bold_font
    else:
        for cell in ws["290:290"]:
            cell.font = bold_font
        
    for rows in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=30):
        for cell in rows:
            cell.fill = PatternFill(fgColor="9FBA78", fill_type = "solid")

    for rows in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=30):
        for cell in rows:
            cell.fill = PatternFill(fgColor="7081db", fill_type = "solid")
        
    
        

    

    # border = Border(left=Side(border_style='thin', color='000000'),
    #     right=Side(border_style='thin', color='000000'),
    #     top=Side(border_style='thin', color='000000'),
    #     bottom=Side(border_style='thin', color='000000'))
    # for rows in ws.iter_rows(min_row=1, max_row=len(salary_slips)+3, min_col=1, max_col=len(salary_slips)+100):

    #     for cell in rows:
    #         cell.border = border


    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'



