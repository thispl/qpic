# Copyright (c) 2021, TEAMPRO and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils import nowdate,nowtime
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook


import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types

@frappe.whitelist()
def download():
    filename = 'WPS Report'
    test = build_xlsx_response(filename)

def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
         
    ws = wb.create_sheet(sheet_name, 0)
    if args.company:
        frappe.errprint(args.company)

        salary_slip = frappe.get_all("Salary Slip",{'company':args.company},['*'])
        if salary_slip:
            frappe.log_error(title='123',message = salary_slip)
            total_salary = frappe.db.get_value('Salary Slip',{'company':args.company,'start_date':args.from_date},['sum(net_pay)'])
            count = frappe.db.count('Salary Slip',{'company':args.company,'start_date':args.from_date})
            employer_id = frappe.db.get_value('Company',{'name':args.company},['employer_id'])
            date = nowdate().replace('-','')
            time = nowtime().replace(':','')
            
            # now_time= time.replaec(microsecond=0)
            ws.append(['Employer ID','File Creation Date','File Creation Time','Payer EID','Payer QID','Payer Bank Short Name','Payer IBAN','Salary Year and Month','Total Salaries','Total Records','SIF Version'])
            ws.append([employer_id,date,time,employer_id,'','CBQ','QA76CBQA000000004020436181001',args.from_date,total_salary,count,''])
            ws.append(['Record Sequence','Employee QID','Employee Visa ID','Employee Name','Employee BAnk Short Name','Employee Account','Salary Frequency','No of Working Days','Net Pay','Basic Salary','Extra Hours','Extra Income','Deductions','Payment Type','Notes / Comments',])
            i = 1
            for s in salary_slip :
                name = s.employee_name
                qid = frappe.get_value('Employee',{'employee':s.employee},['resident_id_number']) or 0
                visa_no = frappe.get_value('Employee',{'employee':s.employee},['visa_application_number']) or 0
                bank_name = s.bank_name
                acc_no = s.bank_account_no
                freq = s.payroll_frequency
                working_days = s.total_working_days
                net_salary = s.net_pay
                basic_salary = s.gross_pay
                extra_inc = 0
                comment = s.comment or ''
                not_amount = frappe.get_value('Additional Salary',{'employee':s.employee,'salary_component':'NOT Amount','payroll_date':s.start_date},['amount']) or 0
                hot_amount = frappe.get_value('Additional Salary',{'employee':s.employee,'salary_component':'HOT Amount','payroll_date':s.start_date},['amount']) or 0
                ot_amount = not_amount + hot_amount
                extra_hrs= s.overtime + s.weekend_ot+ s.holiday_ot

                deductions = s.total_deduction
                ot_hrs = frappe.get_value('Timesheet',{'employee':s.employee},['total_hours']) or 0
                hra = frappe.get_value('Salary Detail',{'salary_component':'House Rental Allowance','parent':s.name },['amount']) or 0
                ot = frappe.get_value('Salary Detail',{'salary_component':'Overtime','parent':s.name },['amount']) or 0
                wot = frappe.get_value('Salary Detail',{'salary_component':'Weekend OT','parent':s.name },['amount']) or 0
                hot = frappe.get_value('Salary Detail',{'salary_component':'Holiday OT','parent':s.name },['amount']) or 0
                ot_cal = ot+wot+hot
                ws.append([i,qid,visa_no,name,bank_name,acc_no,freq,working_days,net_salary,basic_salary,extra_hrs,ot_cal,deductions,'Normal Payment',comment])
                i = i+1
        else:
            company = frappe.get_all("Company")
            for c in company:
                com = c.name
                salary_slip = frappe.get_all("Salary Slip",{'company':com},['*'])
                if salary_slip:
                    frappe.log_error(title='abc',message = salary_slip)
                    total_salary = frappe.db.get_value('Salary Slip',{'company':com,'start_date':args.from_date},['sum(net_pay)'])
                    count = frappe.db.count('Salary Slip',{'company':com,'start_date':args.from_date})
                    employer_id = frappe.db.get_value('Company',{'name':com},['employer_id'])
                    date = nowdate()
                    time = nowtime()
                    ws.append(['Employer ID','File Creation Date','File Creation Time','Payer EID','Payer QID','Payer Bank Short Name','Payer IBAN','Salary Year and Month','Total Salaries','Total Records','SIF Version'])
                    ws.append([employer_id,date,time,employer_id,'','CBQ','QA76CBQA000000004020436181001',args.from_date,total_salary,count,''])
                    ws.append(['Record Sequence','Employee QID','Employee Visa ID','Employee Name','Employee BAnk Short Name','Employee Account','Salary Frequency','No of Working Days','Net Pay','Basic Salary','Extra Income','Deductions','Payment Type','Notes / Comments',])
                    i = 1
                    for s in salary_slip :
                        name = s.employee_name
                        qid = frappe.get_value('Employee',{'employee':s.employee},['resident_id_number']) or 0
                        visa_no = frappe.get_value('Employee',{'employee':s.employee},['visa_application_number']) or 0
                        bank_name = s.bank_name
                        acc_no = s.bank_account_no
                        freq = s.payroll_frequency
                        working_days = s.total_working_days
                        net_salary = s.net_pay
                        basic_salary = s.gross_pay
                        extra_inc = 0
                        comment = s.comment or ''
                        not_amount = frappe.get_value('Additional Salary',{'employee':s.employee,'salary_component':'NOT Amount','payroll_date':s.start_date},['amount']) or 0
                        hot_amount = frappe.get_value('Additional Salary',{'employee':s.employee,'salary_component':'HOT Amount','payroll_date':s.start_date},['amount']) or 0
                        ot_amount = not_amount + hot_amount
                        deductions = s.total_deduction
                        ot_hrs = frappe.get_value('Salary Detail',{'salary_component':'Overtime','parent':s.name },['amount']) or 0
                        hra = frappe.get_value('Salary Detail',{'salary_component':'House Rental Allowance','parent':s.name },['amount']) or 0
                        ot = frappe.get_value('Salary Detail',{'salary_component':'Overtime','parent':s.name },['amount']) or 0
                        wot = frappe.get_value('Salary Detail',{'salary_component':'Weekend OT','parent':s.name },['amount']) or 0
                        hot = frappe.get_value('Salary Detail',{'salary_component':'Holiday OT','parent':s.name },['amount']) or 0
                        ot_cal = ot+wot+hot
                        ws.append([i,qid,visa_no,name,bank_name,acc_no,freq,working_days,net_salary,basic_salary,extra_hrs,ot_cal,deductions,'Normal Payment',comment])
                        i = i+1
        
    
    # company = frappe.get_all("Company")
    # for c in company:
    #     com = c.name
    #     salary_slip = frappe.get_all("Salary Slip",{'company':com},['*'])
    #     if salary_slip:
    #          ws.append(['','','','','','','','','','','',''])
    #          ws.append(['','company',com,'','','','','',''])
    #          ws.append(['','Month','','','','','',''])
    #          ws.append(['','Overtime','','','','','','',''])
    #          ws.append(['SL NO','Id','Name','','Basic','Allowance','HRA','Total Salary','NOT Hrs','HOT Hrs','NOT Amount','HOT Amount','Other Addition','Abs','Abs Deduction','GROSS SALARY','Loan Deduction','Adv Other Deduction','Mess Advance','Net Pay'])
    #     i =1
    #     for s in salary_slip:
    #         div = s.company
    #         emp_id = s.employee
    #         name = s.employee_name
    #         basic = frappe.get_value('Salary Detail',{'salary_component':'Basic','parent':s.name },['amount']) or 0
    #         hra = frappe.get_value('Salary Detail',{'salary_component':'House Rental Allowance','parent':s.name },['amount']) or 0
    #         other_alw = frappe.get_value('Salary Detail',{'salary_component':'Other Allowance','parent':s.name },['amount']) or 0
    #         not_amt = frappe.get_value('Salary Detail',{'salary_component':'NOT Amount','parent':s.name },['amount']) or 0
    #         hot_amt = frappe.get_value('Salary Detail',{'salary_component':'HOT Amount','parent':s.name },['amount']) or 0
    #         others = frappe.get_value('Salary Detail',{'salary_component':'Others','parent':s.name },['amount']) or 0
    #         ab = frappe.get_value('Salary Detail',{'salary_component':'Abs','parent':s.name },['amount']) or 0
    #         abs_deduct = frappe.get_value('Salary Detail',{'salary_component':'Abs Deduction','parent':s.name },['amount']) or 0
    #         gross_pay= s.gross_pay
    #         loan = frappe.get_value('Salary Detail',{'salary_component':'Loan Deduction','parent':s.name },['amount']) or 0
    #         other_deduct = frappe.get_value('Salary Detail',{'salary_component':'Other Deductions','parent':s.name },['amount']) or 0
    #         mess_adv = frappe.get_value('Salary Detail',{'salary_component':'Mess Advance','parent':s.name },['amount']) or 0
    #         net_pay = s.net_pay
    #         ws.append([i,emp_id,name,'',basic,other_alw,hra,'','','',not_amt,hot_amt,others,ab,abs_deduct,gross_pay,loan,other_deduct,mess_adv,net_pay])
    #         i=i+1
  

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'